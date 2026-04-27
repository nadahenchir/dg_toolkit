from app.db.connection import get_connection, get_cursor
from openpyxl import load_workbook

EXCEL_PATH = r'C:\Users\USER\OneDrive\Bureau\dg_toolkit\data\dg toolkit seed data (2).xlsx'

# ------------------------------------------------------------
# Maturity level thresholds — used to resolve raw_score to L1-L5
# Hardcoded here as reference, not stored in DB
# ------------------------------------------------------------
# L1: 0.0        (exactly)
# L2: 0.01–0.49
# L3: 0.50–0.74
# L4: 0.75–0.99
# L5: 1.0        (exactly)

# ------------------------------------------------------------
# Domain weights — 11 domains, equal weight for now
# Adjust if KPMG specifies different weights
# ------------------------------------------------------------
DOMAIN_WEIGHTS = {
    1:  0.0909,
    2:  0.0909,
    3:  0.0909,
    4:  0.0909,
    5:  0.0909,
    6:  0.0909,
    7:  0.0909,
    8:  0.0909,
    9:  0.0909,
    10: 0.0909,
    11: 0.0910,  # slightly higher to ensure sum = 1.0
}

# KPI weights within each domain — equal weight per KPI within domain
# Computed dynamically from KPI count per domain in seed_kpis()

# Question weights — fixed across all questions
QUESTION_WEIGHTS = {
    1: 0.1805,
    2: 0.2071,
    3: 0.2679,
    4: 0.3445,
}

# KPIs where lower raw value = higher maturity
INVERTED_KPI_NAMES = {
    "Master Data Duplicate Rate",
    "Data Integration Error Rate",
}


# ------------------------------------------------------------
# Helpers
# ------------------------------------------------------------

def clean(val):
    """Strip whitespace from string values, return None for empty."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def match_kpi(kpi_name_from_excel, kpi_map):
    """
    Match a KPI name from Excel to a KPI ID in the DB.
    Three-stage matching:
      1. Exact match (normalized)
      2. Substring match (one contains the other)
      3. Word overlap fallback (at least 3 words in common)
    Returns (kpi_id, matched_name) or (None, None).
    """
    n2 = kpi_name_from_excel.lower().strip()

    # Stage 1 — exact
    for name, kid in kpi_map.items():
        if name.lower().strip() == n2:
            return kid, name

    # Stage 2 — substring
    for name, kid in kpi_map.items():
        n1 = name.lower().strip()
        if n1 in n2 or n2 in n1:
            return kid, name

    # Stage 3 — word overlap (at least 3 words in common)
    kpi_words = set(n2.split())
    best_kid, best_name, best_overlap = None, None, 0
    for name, kid in kpi_map.items():
        name_words = set(name.lower().strip().split())
        overlap = len(kpi_words & name_words)
        if overlap >= 3 and overlap > best_overlap:
            best_kid, best_name, best_overlap = kid, name, overlap

    return best_kid, best_name


def load_workbook_sheets():
    print(f"Loading Excel file: {EXCEL_PATH}")
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    print(f"Sheets found: {wb.sheetnames}")
    return wb


# ------------------------------------------------------------
# Step 1 — Seed domains
# ------------------------------------------------------------

def seed_domains(cur):
    print("\n--- Seeding domains ---")

    wb = load_workbook_sheets()
    ws = wb['KPI Catalog']

    domains = []
    display_order = 0

    for row in ws.iter_rows(values_only=True):
        cell = clean(row[1])
        if cell and 'Domain' in cell and cell != 'Domain':
            display_order += 1
            # Extract domain number from "Domain 1: Data Governance"
            parts = cell.split(':')
            domain_num = int(parts[0].replace('Domain', '').strip())
            domain_name = parts[1].strip()
            domains.append((domain_num, domain_name, display_order))

    print(f"Found {len(domains)} domains")

    for domain_id, name, order in domains:
        weight = DOMAIN_WEIGHTS[domain_id]
        cur.execute("""
            INSERT INTO dg_toolkit.domains (id, name, weight, display_order)
            VALUES (%s, %s, %s, %s)
            ON CONFLICT (id) DO UPDATE
                SET name = EXCLUDED.name,
                    weight = EXCLUDED.weight,
                    display_order = EXCLUDED.display_order
        """, (domain_id, name, weight, order))
        print(f"  Domain {domain_id}: {name} (weight={weight})")

    print(f"Domains seeded: {len(domains)}")


# ------------------------------------------------------------
# Step 2 — Seed KPIs
# ------------------------------------------------------------

def seed_kpis(cur):
    print("\n--- Seeding KPIs ---")

    wb = load_workbook_sheets()
    ws = wb['KPI Catalog']

    kpis = []
    current_domain_id = None
    domain_kpi_count = {}  # domain_id -> list of KPI names
    kpi_id = 0

    for row in ws.iter_rows(values_only=True):
        domain_cell = clean(row[1])
        kpi_cell = clean(row[2])
        def_cell = clean(row[3])

        if domain_cell and 'Domain' in domain_cell and domain_cell != 'Domain':
            parts = domain_cell.split(':')
            current_domain_id = int(parts[0].replace('Domain', '').strip())
            if current_domain_id not in domain_kpi_count:
                domain_kpi_count[current_domain_id] = []

        if kpi_cell and kpi_cell != 'KPI Name' and current_domain_id:
            kpi_id += 1
            domain_kpi_count[current_domain_id].append(kpi_id)
            kpis.append({
                'id':        kpi_id,
                'domain_id': current_domain_id,
                'name':      kpi_cell,
                'definition': def_cell,
            })

    print(f"Found {len(kpis)} KPIs")

    # Compute per-domain KPI order and equal weights
    domain_kpi_order = {}  # kpi_id -> (domain_order, weight)
    for domain_id, kpi_ids in domain_kpi_count.items():
        count = len(kpi_ids)
        for i, kid in enumerate(kpi_ids):
            order = i + 1
            # Equal weights within domain, last KPI gets remainder
            if i < count - 1:
                weight = round(1.0 / count, 4)
            else:
                weight = round(1.0 - (round(1.0 / count, 4) * (count - 1)), 4)
            domain_kpi_order[kid] = (order, weight)

    for kpi in kpis:
        kid = kpi['id']
        order, weight = domain_kpi_order[kid]
        is_inverted = kpi['name'] in INVERTED_KPI_NAMES

        cur.execute("""
            INSERT INTO dg_toolkit.kpis
                (id, domain_id, name, definition, domain_order, weight, is_inverted)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (id) DO UPDATE
                SET domain_id    = EXCLUDED.domain_id,
                    name         = EXCLUDED.name,
                    definition   = EXCLUDED.definition,
                    domain_order = EXCLUDED.domain_order,
                    weight       = EXCLUDED.weight,
                    is_inverted  = EXCLUDED.is_inverted
        """, (kid, kpi['domain_id'], kpi['name'], kpi['definition'], order, weight, is_inverted))
        print(f"  KPI {kid}: [{kpi['domain_id']}] {kpi['name']} (weight={weight}, inverted={is_inverted})")

    print(f"KPIs seeded: {len(kpis)}")


# ------------------------------------------------------------
# Step 3 — Seed questions
# ------------------------------------------------------------

def seed_questions(cur):
    print("\n--- Seeding questions ---")

    wb = load_workbook_sheets()
    ws = wb['Question Bank v2']

    # Build KPI name -> ID map from DB
    cur.execute("SELECT id, name FROM dg_toolkit.kpis ORDER BY id")
    kpi_map = {row['name']: row['id'] for row in cur.fetchall()}

    questions = []
    current_kpi_id = None

    rows = list(ws.iter_rows(values_only=True))

    i = 0
    while i < len(rows):
        row = rows[i]

        # Detect KPI header row — column 2 has KPI name, columns 3-6 have Q1-Q4 text
        kpi_cell = clean(row[2])
        q1_text  = clean(row[3])
        q2_text  = clean(row[4])
        q3_text  = clean(row[5])
        q4_text  = clean(row[6])

        if kpi_cell and kpi_cell.startswith('KPI') and q1_text and q1_text.startswith('Q'):
            # This is a KPI header row with question texts
            kpi_name = ':'.join(kpi_cell.split(':')[1:]).strip() if ':' in kpi_cell else kpi_cell
            current_kpi_id, matched_name = match_kpi(kpi_name, kpi_map)

            if current_kpi_id is None:
                print(f"  WARNING: Could not match KPI: '{kpi_cell}'")
                i += 1
                continue
            else:
                print(f"  Matched '{kpi_name}' → KPI {current_kpi_id}: {matched_name}")

            # Extract question texts (strip "Q1: " prefix)
            q_texts = []
            for qt in [q1_text, q2_text, q3_text, q4_text]:
                if qt:
                    text = qt.split(':', 1)[1].strip() if ':' in qt else qt
                    q_texts.append(text)

            if len(q_texts) != 4:
                print(f"  WARNING: Expected 4 questions for KPI {current_kpi_id}, got {len(q_texts)}")
                i += 1
                continue

            # Next 5 rows are the answer option texts (Fully, Mostly, Partially, Slightly, Not)
            opt_rows = []
            j = i + 1
            while j < len(rows) and len(opt_rows) < 5:
                opt_row = rows[j]
                if any(clean(opt_row[k]) for k in range(3, 7)):
                    opt_rows.append([clean(opt_row[k]) for k in range(3, 7)])
                j += 1

            if len(opt_rows) < 5:
                print(f"  WARNING: Could not find 5 option rows for KPI {current_kpi_id}")
                i += 1
                continue

            # opt_rows[0] = Fully, [1] = Mostly, [2] = Partially, [3] = Slightly, [4] = Not
            for q_num in range(1, 5):
                idx = q_num - 1
                questions.append({
                    'kpi_id':          current_kpi_id,
                    'question_number': q_num,
                    'question_text':   q_texts[idx],
                    'weight':          QUESTION_WEIGHTS[q_num],
                    'is_gatekeeper':   q_num == 1,
                    'allows_na':       True,
                    'opt_fully':       opt_rows[0][idx] or '',
                    'opt_mostly':      opt_rows[1][idx] or '',
                    'opt_partially':   opt_rows[2][idx] or '',
                    'opt_slightly':    opt_rows[3][idx] or '',
                    'opt_not':         opt_rows[4][idx] or '',
                })

        i += 1

    print(f"Found {len(questions)} questions")

    for q in questions:
        cur.execute("""
            INSERT INTO dg_toolkit.questions
                (kpi_id, question_number, question_text, weight,
                 is_gatekeeper, allows_na,
                 opt_fully_text, opt_mostly_text, opt_partially_text,
                 opt_slightly_text, opt_not_text)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (kpi_id, question_number) DO UPDATE
                SET question_text    = EXCLUDED.question_text,
                    weight           = EXCLUDED.weight,
                    is_gatekeeper    = EXCLUDED.is_gatekeeper,
                    allows_na        = EXCLUDED.allows_na,
                    opt_fully_text   = EXCLUDED.opt_fully_text,
                    opt_mostly_text  = EXCLUDED.opt_mostly_text,
                    opt_partially_text = EXCLUDED.opt_partially_text,
                    opt_slightly_text  = EXCLUDED.opt_slightly_text,
                    opt_not_text     = EXCLUDED.opt_not_text
        """, (
            q['kpi_id'], q['question_number'], q['question_text'], q['weight'],
            q['is_gatekeeper'], q['allows_na'],
            q['opt_fully'], q['opt_mostly'], q['opt_partially'],
            q['opt_slightly'], q['opt_not']
        ))

    print(f"Questions seeded: {len(questions)}")


# ------------------------------------------------------------
# Step 4 — Seed action library
# ------------------------------------------------------------

def seed_action_library(cur):
    print("\n--- Seeding action library ---")

    wb = load_workbook_sheets()
    ws = wb['ActionLib v2']

    # Build KPI name -> ID map from DB
    cur.execute("SELECT id, name FROM dg_toolkit.kpis ORDER BY id")
    kpi_map = {row['name']: row['id'] for row in cur.fetchall()}

    actions = []
    current_kpi_id = None

    for row in ws.iter_rows(values_only=True):
        kpi_cell    = clean(row[2])
        level_cell  = clean(row[3])
        action_cell = clean(row[4])
        impact_cell = clean(row[5])
        effort_cell = clean(row[6])

        # Detect KPI header
        if kpi_cell and kpi_cell.startswith('KPI'):
            kpi_name = ':'.join(kpi_cell.split(':')[1:]).strip() if ':' in kpi_cell else kpi_cell
            current_kpi_id, matched_name = match_kpi(kpi_name, kpi_map)
            if current_kpi_id is None:
                print(f"  WARNING: Could not match KPI: '{kpi_cell}'")
            else:
                print(f"  Matched '{kpi_name}' → KPI {current_kpi_id}: {matched_name}")

        # Detect action row — level cell looks like "1 → 2"
        if level_cell and '→' in level_cell and action_cell and current_kpi_id:
            try:
                from_level = int(level_cell.split('→')[0].strip())
                actions.append({
                    'kpi_id':     current_kpi_id,
                    'from_level': from_level,
                    'action_text': action_cell,
                    'impact':     impact_cell or 'Medium',
                    'effort':     effort_cell or 'Medium',
                })
            except ValueError:
                print(f"  WARNING: Could not parse level: {level_cell}")

    print(f"Found {len(actions)} actions")

    for a in actions:
        cur.execute("""
            INSERT INTO dg_toolkit.action_library
                (kpi_id, from_level, action_text, impact, effort)
            VALUES (%s, %s, %s, %s, %s)
            ON CONFLICT (kpi_id, from_level) DO UPDATE
                SET action_text = EXCLUDED.action_text,
                    impact      = EXCLUDED.impact,
                    effort      = EXCLUDED.effort
        """, (a['kpi_id'], a['from_level'], a['action_text'], a['impact'], a['effort']))
        print(f"  KPI {a['kpi_id']} L{a['from_level']}→{a['from_level']+1}: {a['action_text'][:60]}...")

    print(f"Actions seeded: {len(actions)}")


# ------------------------------------------------------------
# Main
# ------------------------------------------------------------

def run():
    print("=== DG Toolkit Seed Script ===")
    conn = get_connection()
    cur  = get_cursor(conn)

    try:
        seed_domains(cur)
        conn.commit()
        print("✓ Domains committed")

        seed_kpis(cur)
        conn.commit()
        print("✓ KPIs committed")

        seed_questions(cur)
        conn.commit()
        print("✓ Questions committed")

        seed_action_library(cur)
        conn.commit()
        print("✓ Action library committed")

        print("\n=== Seed completed successfully ===")

        # Final counts
        for table in ['domains', 'kpis', 'questions', 'action_library']:
            cur.execute(f"SELECT COUNT(*) as count FROM dg_toolkit.{table}")
            row = cur.fetchone()
            print(f"  {table}: {row['count']} rows")

    except Exception as e:
        conn.rollback()
        print(f"\n ERROR: {e}")
        raise
    finally:
        cur.close()
        conn.close()


if __name__ == '__main__':
    run()